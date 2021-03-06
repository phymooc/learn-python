#! python3
# updateProduce.py - Correct costs in produce sales spreadsheet

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']


# dictionary
PRICE_UPDATES = {
    'Galic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updateProduceSales.xlsx')
